from decimal import Decimal
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class BudgetExportService:
    @staticmethod
    async def export_to_excel(data: list[dict]) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Аналитический отчет"

        # Определяем колонки (ключ в словаре -> заголовок в Excel)
        columns = [
            ("kcsr_code", "Код ЦСР", 20),
            ("kcsr_name", "Наименование ЦСР", 40),
            ("budget_name", "Бюджет", 25),
            ("budget_period", "Период", 12),
            ("limit_pbs", "Лимит ПБС", 18),
            ("rcb_payments", "Исполнение (РКБ)", 18),
            ("buau_payments", "Исполнение (БУАУ)", 18),
            ("agr_amount", "Сумма соглашений", 18),
            ("gz_contracts_amount", "Сумма контрактов", 18),
            ("gz_paid", "Оплачено (ГЗ)", 18),
        ]

        # Стилизация
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Рисуем шапку
        for col_idx, (key, title, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Заполнятельное данных
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, (key, title, width) in enumerate(columns, 1):
                value = row_data.get(key)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                # Форматирование денежных колонок
                if isinstance(value, (int, float, Decimal)):
                    cell.number_format = "#,##0.00"
                    cell.alignment = right_align
                else:
                    cell.alignment = Alignment(vertical="center")

        # Сохранение в буфер памяти
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
